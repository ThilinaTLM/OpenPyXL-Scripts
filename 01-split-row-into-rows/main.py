#! /usr/bin/python3

from openpyxl import Workbook, load_workbook
from os import path
from sys import argv

# Configurations ------------------------------------------------------------------------------------------------------------
COMMON_PART_LEN = 11
COMMON_PART_HEADINGS = [
    'Code',
    'Name',
    'Total investments',
    'Number investments',
    'No investments',
    'Portfolio size',
    'Number exits',
    'New investments',
    'Lead investor',
    'Lead reason',
    'Email'
]

DATA_ENTRY_PART_LEN = 7
DATA_ENTRY_PART_HEADINGS = [
    'ID_1',
    'Capital_Invested_1',
    'Round_Size_1',
    'Share_1',
    'Sector_1',
    'Syndicate_1',
    'Round_Numbber_1'
]

START_ROW_NUMBER = 2
END_ROW_NUMBER = None
START_SHEET_NUMBER = 0
END_SHEET_NUMBER = None

COMMON_PART_ONLY = True
FILES = ['Source.xlsx']
# ---------------------------------------------------------------------------------------------------------------------------

def log(topic, msg, extra=""):
    """
        Used to Console Log
    """
    if not extra == "":
        extra = f'({extra}):'
    print(f'[{topic}]:', msg, extra)


def assignRow(sheet, data, row_number):
    """
        Assign list of data into one row
    """
    for i, d in enumerate(data):
        sheet.cell(row=row_number, column=i + 1).value = d


def appendRow(sheet, data):
    """
        Append a new row into a sheet and assign list of data
    """
    assignRow(sheet, data, sheet.max_row + 1)


def observeRow(row):
    """
        Spliting a row into common part and sub parts
    """
    ROW_LENGTH = len(row)
    point = 11

    common = row[0:COMMON_PART_LEN]  # Common Part
    parts = []  # Data Entry Parts

    while point < ROW_LENGTH:
        sub_entry = row[point:point + DATA_ENTRY_PART_LEN]
        if any(sub_entry):
            parts.append(sub_entry)
        point += DATA_ENTRY_PART_LEN

    return [common, parts]


def refactoringSheet(sourceSheet, outputSheet):
    """
        Refactoring sourceSheet into outputSheet
    """
    global END_ROW_NUMBER
    if END_ROW_NUMBER == None:
        END_ROW_NUMBER = sourceSheet.max_row
    for ind, row in enumerate(
            sourceSheet.iter_rows(min_row=START_ROW_NUMBER, max_row=END_ROW_NUMBER, values_only=True)):
        log('SH', 'Reading row, ', ind + START_ROW_NUMBER)

        common, parts = observeRow(row)
        if COMMON_PART_ONLY and not parts:
            log('SH', 'Common part only is enabled')
            appendRow(outputSheet, common)

        print('[SH]:', 'row', ind + START_ROW_NUMBER, 'is refactored into', len(parts), 'rows')

        for p in parts:
            appendRow(outputSheet, common + p)


def refactorWorkbook(path, output_path):
    log('FILE', 'Opening source file....', path)
    wb_source = load_workbook(path, read_only=True)
    log('FILE', 'Opened', path)

    if END_SHEET_NUMBER == None:
        log('WB', 'If mutiple sheets are exists, all will be refactored')
        sheets = wb_source.worksheets[START_SHEET_NUMBER - 1:]
    else:
        sheets = wb_source.worksheets[START_SHEET_NUMBER - 1:END_SHEET_NUMBER + 1]

    wb_output = Workbook()

    if len(sheets) == 0:
        log('WB', 'No Sheets found, nothing to process', path)
    else:
        for source_sheet in sheets:
            log('SH', 'Refactoring sheet', path + " > " + source_sheet.title)
            wb_output.create_sheet(source_sheet.title)
            output_sheet = wb_output.worksheets[-1]

            log('SH', 'Setting column headers', path + " > " + source_sheet.title)
            assignRow(output_sheet, COMMON_PART_HEADINGS + DATA_ENTRY_PART_HEADINGS, 1)

            refactoringSheet(source_sheet, output_sheet)

        wb_output.remove(wb_output.worksheets[0])
        wb_output.save(output_path)
        log('WB', "Refactored workbook is saved!", output_path)


def getFilePath():
    log('FILE', 'Enter source file path', '(relative / absolute)')
    source = ""
    while not source:
        source = input('SOURCE FILE: ').strip()

    directory, filename = path.split(source)
    default_output_name = "_output_" + filename

    default_output = path.join(directory, default_output_name)

    log('FILE', 'Enter output file path', '(relative / absolute)')
    output = input('OUTPUT FILE (' + default_output_name + '): ').strip()

    if output == "":
        output = default_output

    return [source, output]


def getFilePath_fromArgs():
    if len(argv) > 1:
        source = argv[1]
        log('FILE', 'Taking source file path from CLI arguments', source)
    else:
        return [None, None]

    if len(argv) > 2:
        output = argv[2]
        log('FILE', 'Taking output file path from CLI arguments', output)
    else:
        directory, filename = path.split(source)
        default_output_name = "_output_" + filename
        output = path.join(directory, default_output_name)

        log('FILE', 'Use default output file format', output)

    return [source, output]

def generateOutputPath(source):
    directory, filename = path.split(source)
    default_output_name = "_output_" + filename

    return path.join(directory, default_output_name)

def main():
    source, output = getFilePath()
    refactorWorkbook(source, output)


if len(argv) > 1:
    source, output = getFilePath_fromArgs()
    refactorWorkbook(source, output)
    exit()

if FILES:
    for source in FILES:
        output = generateOutputPath(source)
        log('FILE', 'Source file', source)
        log('FILE', 'Output file', output)
        refactorWorkbook(source, output)
    exit()

main()
while True:
    more = input('Do you have more work ? (y/N): ').lower()
    if more == "" or more == "n":
        exit()
    elif more == "y":
        main()
